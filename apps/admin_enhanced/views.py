"""
Export views for admin changelist.
"""
import json
from django.http import HttpResponse, JsonResponse
from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.translation import gettext_lazy as _, gettext
from django.db import models
from django.template.loader import render_to_string
from django.shortcuts import get_object_or_404


@staff_member_required
def export_changelist(request, app_label, model_name, format_type):
    """
    Export changelist data to Excel or JSON.
    
    Args:
        app_label: The app label of the model
        model_name: The model name
        format_type: 'excel' or 'json'
    """
    try:
        # Get the model from Django's app registry
        from django.apps import apps
        try:
            model = apps.get_model(app_label, model_name)
        except LookupError:
            return JsonResponse({'error': _('Model not found')}, status=404)
        
        # Get the ModelAdmin instance
        if model not in admin.site._registry:
            return JsonResponse({'error': _('ModelAdmin not found')}, status=404)
        
        model_admin = admin.site._registry[model]
        
        # Get the queryset using ModelAdmin's get_queryset method
        # This respects the ModelAdmin's custom queryset filtering
        queryset = model_admin.get_queryset(request)
        
        # Apply search filter if present
        search_term = request.GET.get('q', '').strip()
        if search_term and model_admin.search_fields:
            from django.db.models import Q
            search_q = Q()
            for field_name in model_admin.search_fields:
                if '__' in field_name:
                    # Related field search
                    search_q |= Q(**{f"{field_name}__icontains": search_term})
                else:
                    # Direct field search
                    search_q |= Q(**{f"{field_name}__icontains": search_term})
            queryset = queryset.filter(search_q)
        
        # Apply list filters from request parameters
        # Django admin filters are passed as query parameters
        for key, value in request.GET.items():
            if key in ['q', 'page', 'o', '_changelist_filters', 'columns[]', 'selected_ids[]']:
                continue
            if value:
                try:
                    # Try to apply the filter
                    if '__' in key:
                        queryset = queryset.filter(**{key: value})
                    elif hasattr(model, key):
                        queryset = queryset.filter(**{key: value})
                except Exception:
                    # Skip invalid filters
                    pass
        
        # Get selected rows (if any)
        selected_ids = request.GET.getlist('selected_ids[]')
        if selected_ids:
            queryset = queryset.filter(pk__in=selected_ids)
        
        # Get selected columns
        selected_columns = request.GET.getlist('columns[]')
        if not selected_columns:
            # Use list_display if no columns specified
            selected_columns = list(model_admin.list_display) if model_admin.list_display else []
        
        # Remove action_checkbox and action_buttons if present
        selected_columns = [col for col in selected_columns if col not in ['action_checkbox', 'action_buttons']]
        
        # Get field labels (ensure all are strings, not translation proxies)
        from django.utils.encoding import force_str
        field_labels = {}
        for field_name in selected_columns:
            if hasattr(model_admin, field_name):
                # Custom method
                method = getattr(model_admin, field_name)
                label = getattr(method, 'short_description', None)
                if label:
                    field_labels[field_name] = force_str(label)
                else:
                    field_labels[field_name] = field_name.replace('_', ' ').title()
            elif hasattr(model, field_name):
                # Model field
                try:
                    field = model._meta.get_field(field_name)
                    if field.verbose_name:
                        field_labels[field_name] = force_str(field.verbose_name)
                    else:
                        field_labels[field_name] = field_name.replace('_', ' ').title()
                except Exception:
                    field_labels[field_name] = field_name.replace('_', ' ').title()
            else:
                field_labels[field_name] = field_name.replace('_', ' ').title()
        
        # Prepare data
        from django.utils.encoding import force_str
        data_rows = []
        for obj in queryset:
            row = {}
            for field_name in selected_columns:
                try:
                    if hasattr(model_admin, field_name):
                        # Custom method
                        value = getattr(model_admin, field_name)(obj)
                        # Convert to JSON-serializable type
                        if value is None:
                            row[field_name] = None
                        elif isinstance(value, (str, int, float, bool)):
                            row[field_name] = value
                        elif isinstance(value, models.Model):
                            row[field_name] = force_str(value)
                        else:
                            # Convert any other type to string
                            row[field_name] = force_str(value)
                    elif hasattr(obj, field_name):
                        # Model field
                        try:
                            field = model._meta.get_field(field_name)
                        except Exception:
                            field = None
                        
                        value = getattr(obj, field_name)
                        if value is None:
                            row[field_name] = None
                        elif field and isinstance(field, models.ForeignKey):
                            row[field_name] = force_str(value)
                        elif field and isinstance(field, models.ManyToManyField):
                            row[field_name] = ', '.join([force_str(item) for item in value.all()])
                        elif field and isinstance(field, models.DateTimeField):
                            row[field_name] = value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
                        elif field and isinstance(field, models.DateField):
                            row[field_name] = value.strftime('%Y-%m-%d') if value else ''
                        elif isinstance(value, (str, int, float, bool)):
                            row[field_name] = value
                        else:
                            row[field_name] = force_str(value)
                    else:
                        row[field_name] = ''
                except Exception as e:
                    row[field_name] = f'Error: {str(e)}'
            data_rows.append(row)
        
        # Export based on format
        # Ensure model_name is a string, not a translation proxy
        model_name_str = force_str(model._meta.verbose_name_plural)
        if format_type == 'json':
            return export_json(data_rows, field_labels, model_name_str)
        elif format_type == 'excel':
            return export_excel(data_rows, field_labels, model_name_str)
        else:
            return JsonResponse({'error': str(_('Invalid format type'))}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def export_json(data_rows, field_labels, model_name):
    """Export data to JSON format."""
    from django.utils.encoding import force_str
    # Use field labels as keys
    result = []
    for row in data_rows:
        labeled_row = {}
        for field_name, value in row.items():
            label = field_labels.get(field_name, field_name)
            # Ensure label is a string
            label_str = force_str(label) if label else str(field_name)
            # Ensure value is JSON-serializable
            if value is None:
                labeled_row[label_str] = None
            elif isinstance(value, (str, int, float, bool)):
                labeled_row[label_str] = value
            else:
                # Convert to string for any other type
                labeled_row[label_str] = force_str(value)
        result.append(labeled_row)
    
    response = HttpResponse(
        json.dumps(result, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    model_name_str = force_str(model_name) if model_name else 'export'
    response['Content-Disposition'] = f'attachment; filename="{model_name_str}_export.json"'
    return response


def export_excel(data_rows, field_labels, model_name):
    """Export data to Excel format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        # Note: Install openpyxl for Excel export: pip install openpyxl
        return export_csv(data_rows, field_labels, model_name)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(model_name)[:31]  # Excel sheet name limit
    
    # Write headers
    headers = [field_labels.get(field_name, field_name) for field_name in data_rows[0].keys()] if data_rows else []
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, field_name in enumerate(row.keys(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[field_name])
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_idx in range(2, len(data_rows) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export.xlsx"'
    wb.save(response)
    return response


def export_csv(data_rows, field_labels, model_name):
    """Fallback CSV export if openpyxl not available."""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    if data_rows:
        headers = [field_labels.get(field_name, field_name) for field_name in data_rows[0].keys()]
        writer.writerow(headers)
        
        # Write data
        for row in data_rows:
            writer.writerow([row.get(field_name, '') for field_name in data_rows[0].keys()])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export.csv"'
    return response


@staff_member_required
def view_object_modal(request, app_label, model_name, object_id):
    """
    Return readonly HTML content for modal view.
    """
    from django.utils.translation import gettext as _
    
    def return_error_html(error_msg, status_code=500):
        """Return HTML error message instead of JSON"""
        error_html = f'''
        <div class="alert alert-danger">
          <i class="fas fa-exclamation-triangle"></i> 
          <strong>{_("Error")}:</strong> {error_msg}
        </div>
        '''
        return HttpResponse(error_html, status=status_code)
    
    try:
        from django.apps import apps
        model = apps.get_model(app_label, model_name)
    except LookupError as e:
        return return_error_html(f'Model not found: {str(e)}', 404)
    except Exception as e:
        return return_error_html(f'Error loading model: {str(e)}', 500)
    
    # Get the ModelAdmin instance
    if model not in admin.site._registry:
        return return_error_html(_('ModelAdmin not found'), 404)
    
    model_admin = admin.site._registry[model]
    
    try:
        # Get the object
        obj = get_object_or_404(model, pk=object_id)
    except Exception as e:
        return return_error_html(f'Error loading object: {str(e)}', 500)
    
    # Check permissions
    try:
        has_view_permission = getattr(model_admin, 'has_view_permission', None)
        if has_view_permission:
            has_perm = has_view_permission(request, obj)
        else:
            has_perm = model_admin.has_change_permission(request, obj)
        
        if not has_perm:
            return return_error_html(_('Permission denied'), 403)
    except Exception as e:
        return return_error_html(f'Permission check error: {str(e)}', 500)
    
    try:
        # Use a simpler approach: directly use Django admin's changeform_view with readonly mode
        # Create a temporary request with readonly=1 parameter
        from django.test import RequestFactory
        from copy import copy
        
        # Get fieldsets from ModelAdmin
        fieldsets = model_admin.get_fieldsets(request, obj)
        if not fieldsets:
            # If no fieldsets defined, get all fields
            from django.db import models as db_models
            model_fields = []
            for field in model._meta.get_fields():
                if isinstance(field, db_models.Field) and not isinstance(field, db_models.ManyToManyRel):
                    if not isinstance(field, db_models.AutoField):
                        model_fields.append(field.name)
            fieldsets = [(None, {'fields': model_fields})]
        
        # Prepare field values and labels
        field_values = {}
        field_labels = {}
        
        for fieldset_name, fieldset_options in fieldsets:
            for field_name in fieldset_options.get('fields', []):
                if isinstance(field_name, str):
                    try:
                        # Get field value
                        field = model._meta.get_field(field_name)
                        value = getattr(obj, field_name, None)
                        
                        # Format value based on field type
                        from django.db import models as db_models
                        if isinstance(field, db_models.BooleanField):
                            field_values[field_name] = gettext('Yes') if value else gettext('No')
                        elif isinstance(field, db_models.DateTimeField) and value:
                            from django.utils import timezone
                            if timezone.is_aware(value):
                                field_values[field_name] = timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                field_values[field_name] = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(field, db_models.DateField) and value:
                            field_values[field_name] = value.strftime('%Y-%m-%d')
                        elif isinstance(field, db_models.ForeignKey) and value:
                            field_values[field_name] = str(value)
                        elif isinstance(field, db_models.ManyToManyField):
                            field_values[field_name] = ', '.join([str(item) for item in value.all()]) if value else ''
                        elif value is None:
                            field_values[field_name] = '-'
                        else:
                            field_values[field_name] = str(value)
                        
                        # Get field label
                        field_labels[field_name] = field.verbose_name or field_name.replace('_', ' ').title()
                    except Exception as e:
                        field_values[field_name] = f'Error: {str(e)}'
                        field_labels[field_name] = field_name.replace('_', ' ').title()
        
        # Create simplified fieldsets structure for template
        simplified_fieldsets = []
        for fieldset_name, fieldset_options in fieldsets:
            simplified_fieldsets.append({
                'name': fieldset_name,
                'description': fieldset_options.get('description', ''),
                'fields': fieldset_options.get('fields', [])
            })
        
        # Render using a simpler template
        # Convert dict to list of tuples for easier template access
        field_values_list = list(field_values.items())
        field_labels_list = list(field_labels.items())
        
        context = {
            'fieldsets': simplified_fieldsets,
            'field_values': field_values,
            'field_labels': field_labels,
            'field_values_list': field_values_list,
            'field_labels_list': field_labels_list,
            'original': obj,
            'opts': model._meta,
        }
        
        try:
            html = render_to_string('admin/includes/view_modal_simple.html', context, request=request)
            return HttpResponse(html)
        except Exception as template_error:
            # Fallback: try using AdminForm approach
            try:
                from django.contrib.admin.helpers import AdminForm
                from django import forms
                
                class ReadonlyForm(forms.ModelForm):
                    class Meta:
                        model = model
                        fields = '__all__'
                
                form = ReadonlyForm(instance=obj)
                for field_name in form.fields:
                    form.fields[field_name].disabled = True
                
                readonly_fields = list(field_values.keys())
                adminform = AdminForm(form, fieldsets, {}, readonly_fields=readonly_fields, model_admin=model_admin)
                
                context = {
                    'adminform': adminform,
                    'inline_admin_formsets': [],
                    'original': obj,
                    'opts': model._meta,
                    'is_readonly': True,
                }
                
                html = render_to_string('admin/includes/view_modal_content.html', context, request=request)
                return HttpResponse(html)
            except Exception as fallback_error:
                import traceback
                error_detail = traceback.format_exc()
                error_html = f'''
                <div class="alert alert-danger">
                  <i class="fas fa-exclamation-triangle"></i> 
                  <strong>{_("Error")}:</strong> {str(template_error)}
                  <br><small>Fallback also failed: {str(fallback_error)}</small>
                  <details style="margin-top: 10px;">
                    <summary>{_("Error Details")}</summary>
                    <pre style="white-space: pre-wrap; font-size: 0.8em;">{error_detail}</pre>
                  </details>
                </div>
                '''
                return HttpResponse(error_html, status=500)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        from django.utils.translation import gettext as _
        error_html = f'''
        <div class="alert alert-danger">
          <i class="fas fa-exclamation-triangle"></i> 
          <strong>{_("Error")}:</strong> Error rendering view: {str(e)}
          <details style="margin-top: 10px;">
            <summary>{_("Error Details")}</summary>
            <pre style="white-space: pre-wrap; font-size: 0.8em;">{error_detail}</pre>
          </details>
        </div>
        '''
        return HttpResponse(error_html, status=500)
